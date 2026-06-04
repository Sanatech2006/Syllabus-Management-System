import openpyxl
import json
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.http import JsonResponse
from django.urls import reverse
from openpyxl.styles import Font, PatternFill
from .models import CourseStr, CourseContent, normalize_course_code
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from modules.program_manage.models import Program, normalize_program_code, normalize_program_value


UPLOAD_FILTER_FIELDS = (
    'prog_code',
    'branch',
    'year',
    'prog_type',
    'prog_category',
    'sem',
    'course_code',
    'part',
    'course_category',
    'course_title',
    'hrs_per_week',
    'credit',
    'marks_cia',
    'marks_ese',
    'total_marks',
)


def _upload_center_table_redirect():
    return redirect(f"{reverse('upload_center:upload_center')}#saved-courses-section")


def _distinct_non_empty(queryset, field_name):
    queryset = queryset.exclude(**{f'{field_name}__isnull': True})
    field = CourseStr._meta.get_field(field_name)

    if getattr(field, 'empty_strings_allowed', False):
        queryset = queryset.exclude(**{field_name: ''})

    return list(
        queryset.values_list(field_name, flat=True).distinct().order_by(field_name)
    )


def _apply_upload_filters(queryset, filters, exclude_field=None):
    for field in UPLOAD_FILTER_FIELDS:
        if field == exclude_field:
            continue

        value = filters.get(field)
        if value:
            queryset = queryset.filter(**{field: value})

    return queryset


def _build_upload_filter_options(base_queryset, filters):
    option_querysets = {
        field: _apply_upload_filters(base_queryset, filters, exclude_field=field)
        for field in UPLOAD_FILTER_FIELDS
    }

    return {
        'prog_codes': _distinct_non_empty(option_querysets['prog_code'], 'prog_code'),
        'branches': _distinct_non_empty(option_querysets['branch'], 'branch'),
        'years': _distinct_non_empty(option_querysets['year'], 'year'),
        'prog_types': _distinct_non_empty(option_querysets['prog_type'], 'prog_type'),
        'prog_categories': _distinct_non_empty(option_querysets['prog_category'], 'prog_category'),
        'semesters': _distinct_non_empty(option_querysets['sem'], 'sem'),
        'course_codes': _distinct_non_empty(option_querysets['course_code'], 'course_code'),
        'parts': _distinct_non_empty(option_querysets['part'], 'part'),
        'course_categories': _distinct_non_empty(option_querysets['course_category'], 'course_category'),
        'course_titles': _distinct_non_empty(option_querysets['course_title'], 'course_title'),
        'hrs_per_week_options': [str(value).rstrip('0').rstrip('.') for value in _distinct_non_empty(option_querysets['hrs_per_week'], 'hrs_per_week')],
        'credit_options': [str(value).rstrip('0').rstrip('.') for value in _distinct_non_empty(option_querysets['credit'], 'credit')],
        'marks_cia_options': [str(value).rstrip('0').rstrip('.') for value in _distinct_non_empty(option_querysets['marks_cia'], 'marks_cia')],
        'marks_ese_options': [str(value).rstrip('0').rstrip('.') for value in _distinct_non_empty(option_querysets['marks_ese'], 'marks_ese')],
        'total_marks_options': [str(value).rstrip('0').rstrip('.') for value in _distinct_non_empty(option_querysets['total_marks'], 'total_marks')],
    }


def _course_form_initial(course):
    return {
        'prog_code': course.prog_code or '',
        'branch': course.branch or '',
        'degree': course.degree or course.year or '',
        'prog_type': course.prog_type or '',
        'prog_category': course.prog_category or '',
        'sem': course.sem or '',
        'course_code': course.course_code or '',
        'part': course.part or '',
        'course_category': course.course_category or '',
        'course_title': course.course_title or '',
        'hrs_per_week': course.hrs_per_week or '',
        'credit': course.credit or '',
        'marks_cia': course.marks_cia or '',
        'marks_ese': course.marks_ese or '',
        'total_marks': course.total_marks or '',
    }


def _program_match_from_post(post_data):
    return {
        'prog_type': normalize_program_value(post_data.get("prog_type")).upper(),
        'prog_category': normalize_program_value(post_data.get("prog_category")).title(),
        'degree': normalize_program_value(post_data.get("degree")),
        'branch': normalize_program_value(post_data.get("branch")),
        'prog_code': normalize_program_code(post_data.get("prog_code")),
    }


def _program_exists(program_data):
    return Program.objects.filter(is_active=True, **program_data).exists()


def _program_form_context(form_values=None):
    active_programs = Program.objects.filter(is_active=True)

    program_rows = list(
        active_programs
        .values('prog_code', 'branch', 'degree', 'prog_type', 'prog_category')
        .order_by('prog_code', 'branch')
    )
    prog_codes = list(
        active_programs.exclude(prog_code='')
        .values_list('prog_code', flat=True)
        .distinct()
        .order_by('prog_code')
    )

    return {
        "form_values": form_values or {},
        "program_rows_json": json.dumps(program_rows),
        "prog_codes": prog_codes,
    }


@login_required(login_url='/login/')
def upload_center(request):
    filters = {field: request.GET.get(field) for field in UPLOAD_FILTER_FIELDS}
    base_queryset = CourseStr.objects.all()
    courses = _apply_upload_filters(base_queryset, filters).order_by('-created_at')
    per_page_default = 10
    allowed_per_page = {10, 20, 50, 100}
    try:
        per_page = int(request.GET.get('per_page', per_page_default))
        if per_page not in allowed_per_page:
            per_page = per_page_default
    except (ValueError, TypeError):
        per_page = per_page_default

    paginator = Paginator(courses, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Set has_pdf dynamically
    for course in page_obj:
        course.has_pdf = CourseContent.objects.filter(
            course_code=course.course_code,
            pdf__isnull=False
        ).exclude(pdf='').exists()

    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'upload_center.html', {
        'courses': page_obj,
        'page_obj': page_obj,
        'per_page': per_page,
        'pagination_query': query_params.urlencode(),
        **_build_upload_filter_options(base_queryset, filters),
    })


# 🔵 Upload → Blue "Uploaded"
def upload_course_content(request):
    if not request.user.is_authenticated:
        return redirect('/login/')

    # Allow HOD (is_staff) and Admin (is_superuser) to upload
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'You do not have permission to upload files.')
        return _upload_center_table_redirect()

    if request.method == 'POST' and request.FILES.get('pdf_file'):
        course_code = request.POST.get('course_code', '')
        pdf_file = request.FILES['pdf_file']

        safe_code = normalize_course_code(course_code)

        if not safe_code:
            messages.error(request, 'Missing course code.')
            return _upload_center_table_redirect()

        course_content, created = CourseContent.objects.get_or_create(course_code=safe_code)

        if course_content.pdf:
            existing_name = course_content.pdf.name
            target_name = f'course_pdfs/{safe_code}.pdf'
            if existing_name and existing_name != target_name:
                course_content.pdf.delete(save=False)

        try:
            course_content.pdf = pdf_file
            course_content.save()
        except PermissionError:
            messages.error(
                request,
                f"The existing PDF for {safe_code} is currently open somewhere. Close the PDF tab/viewer or Explorer preview, then upload again."
            )
            return _upload_center_table_redirect()

        # RESET STATES
        CourseStr.objects.filter(course_code=safe_code).update(
            is_saved=False,
            is_finalized=False
        )

        messages.success(request, f'PDF queued for {safe_code}.')
        return _upload_center_table_redirect()

    messages.error(request, 'No file uploaded.')
    return _upload_center_table_redirect()


# 🟢 Save → Green "Saved"
def save_courses(request):
    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect('/login/')

        if not request.user.is_superuser:
            messages.error(request, 'Only administrators can save courses.')
            return redirect('upload_center:upload_center')
        courses = CourseStr.objects.all()

        saved_count = 0

        for course in courses:
            has_pdf = CourseContent.objects.filter(
                course_code=course.course_code,
                pdf__isnull=False
            ).exclude(pdf='').exists()

            if has_pdf and not course.is_finalized:
                if not course.is_saved:
                    saved_count += 1
                course.is_saved = True
                course.save()

        if saved_count == 1:
            messages.success(request, "1 PDF uploaded successfully.")
        elif saved_count > 1:
            messages.success(request, f"{saved_count} PDFs uploaded successfully.")
        else:
            messages.info(request, "No queued PDFs to upload.")

    return redirect('upload_center:upload_center')


# ⚪ Save & Confirm → Grey "Uploaded ✔"
def finalize_courses(request):
    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect('/login/')

        if not request.user.is_superuser:
            messages.error(request, 'Only administrators can finalize courses.')
            return redirect('upload_center:upload_center')
        count = CourseStr.objects.filter(is_saved=True).update(is_finalized=True)

        if count > 0:
            messages.success(request, f"{count} courses uploaded successfully.")
        else:
            messages.info(request, "No saved courses to upload.")

    return redirect('upload_center:upload_center')


def delete_course(request, course_id):
    if request.method == 'POST':
        if not request.user.is_authenticated:
            return redirect('/login/')

        # Allow both administrators and HODs (staff) to delete
        if not (request.user.is_superuser or request.user.is_staff):
            messages.error(request, 'You do not have permission to delete courses.')
            return redirect('upload_center:upload_center')

        try:
            course = CourseStr.objects.get(id=course_id)
            course_code = normalize_course_code(course.course_code)

            course.delete()

            if course_code:
                content = CourseContent.objects.filter(course_code=course_code).first()
                if content:
                    if content.pdf:
                        content.pdf.delete(save=False)
                    content.delete()

            messages.success(request, 'Course deleted successfully.')
        except CourseStr.DoesNotExist:
            messages.error(request, 'Course not found.')

    return redirect('upload_center:upload_center')

def delete_pdf(request, course_id):
    if request.method == 'POST':
        if not request.user.is_authenticated:
            return redirect('/login/')

        # Allow both administrators and HODs (staff) to delete PDFs
        if not (request.user.is_superuser or request.user.is_staff):
            messages.error(request, 'You do not have permission to delete syllabi.')
            return redirect('upload_center:upload_center')

        try:
            course = CourseStr.objects.get(id=course_id)
            course_code = normalize_course_code(course.course_code)

            content = CourseContent.objects.filter(course_code=course_code).first()
            if content and content.pdf:
                content.pdf.delete(save=False)
                content.pdf = None
                content.save()

                CourseStr.objects.filter(course_code=course_code).update(
                    is_saved=False,
                    is_finalized=False
                )
                messages.success(request, f'Syllabus PDF deleted for {course_code}.')
            else:
                messages.warning(request, f'No syllabus found for {course_code}.')

        except CourseStr.DoesNotExist:
            messages.error(request, 'Course not found.')

    return redirect('upload_center:upload_center')


def add_course(request):
    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect('/login/')

        # Allow both Admins and HODs (staff) to add courses
        if not (request.user.is_superuser or request.user.is_staff):
            messages.error(request, 'You do not have permission to add courses.')
            return redirect('upload_center:upload_center')
        course_code = normalize_course_code(request.POST.get("course_code"))
        program_data = _program_match_from_post(request.POST)

        if course_code and CourseStr.objects.filter(course_code=course_code).exists():
            messages.error(request, f"Course code {course_code} already exists. Use a different course code.")
            return render(request, "add_course.html", {
                **_program_form_context(request.POST),
                "form_mode": "add",
            })

        if not _program_exists(program_data):
            messages.error(
                request,
                "This program does not exist in Program Management. Add the program first before adding courses."
            )
            return render(request, "add_course.html", {
                **_program_form_context(request.POST),
                "form_mode": "add",
            })

        CourseStr.objects.create(
            degree=program_data["degree"],
            prog_type=program_data["prog_type"],
            prog_category=program_data["prog_category"],
            prog_code=program_data["prog_code"],
            branch=program_data["branch"],
            sem=request.POST.get("sem"),
            course_code=course_code,
            part=request.POST.get("part"),
            course_category=request.POST.get("course_category"),
            course_title=request.POST.get("course_title"),
            hrs_per_week=request.POST.get("hrs_per_week") or 0,
            credit=request.POST.get("credit") or 0,
            marks_cia=request.POST.get("marks_cia") or 0,
            marks_ese=request.POST.get("marks_ese") or 0,
            total_marks=request.POST.get("total_marks") or 0,
            is_saved=False,
            is_finalized=False,
        )

        # If a PDF was uploaded with the add-course form, save it to CourseContent
        if request.FILES.get('pdf_file'):
            pdf_file = request.FILES['pdf_file']
            safe_code = course_code

            if not safe_code:
                messages.error(request, 'Missing course code for uploaded PDF.')
            else:
                course_content, created = CourseContent.objects.get_or_create(course_code=safe_code)

                if course_content.pdf:
                    existing_name = course_content.pdf.name
                    target_name = f'course_pdfs/{safe_code}.pdf'
                    if existing_name and existing_name != target_name:
                        course_content.pdf.delete(save=False)

                try:
                    course_content.pdf = pdf_file
                    course_content.save()
                except PermissionError:
                    messages.error(
                        request,
                        f"The existing PDF for {safe_code} is currently open somewhere. Close the PDF tab/viewer or Explorer preview, then upload again."
                    )
                else:
                    # Reset saved/finalized states for this course
                    CourseStr.objects.filter(course_code=safe_code).update(
                        is_saved=False,
                        is_finalized=False
                    )
                    messages.success(request, f'PDF queued for {safe_code}.')

        if "add_next" in request.POST:
            return redirect('upload_center:add_course')

        return redirect('upload_center:upload_center')

    return render(request, "add_course.html", {
        **_program_form_context(),
        "form_mode": "add",
    })


@login_required(login_url='/login/')
def edit_course(request, course_id):
    course = get_object_or_404(CourseStr, id=course_id)

    if not request.user.is_authenticated:
        return redirect('/login/')

    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can edit courses.')
        return redirect('upload_center:upload_center')

    if request.method == "POST":
        old_course_code = normalize_course_code(course.course_code)
        new_course_code = normalize_course_code(request.POST.get("course_code"))
        program_data = _program_match_from_post(request.POST)

        if (
            new_course_code
            and CourseStr.objects.filter(course_code=new_course_code).exclude(id=course.id).exists()
        ):
            messages.error(request, f"Course code {new_course_code} already exists. Use a different course code.")
            return render(request, "add_course.html", {
                **_program_form_context(request.POST),
                "form_mode": "edit",
                "course": course,
            })

        if not _program_exists(program_data):
            messages.error(
                request,
                "This program does not exist in Program Management. Add the program first before assigning a course to it."
            )
            return render(request, "add_course.html", {
                **_program_form_context(request.POST),
                "form_mode": "edit",
                "course": course,
            })

        course.degree = program_data["degree"]
        course.prog_type = program_data["prog_type"]
        course.prog_category = program_data["prog_category"]
        course.prog_code = program_data["prog_code"]
        course.branch = program_data["branch"]
        course.sem = request.POST.get("sem")
        course.course_code = new_course_code
        course.part = request.POST.get("part")
        course.course_category = request.POST.get("course_category")
        course.course_title = request.POST.get("course_title")
        course.hrs_per_week = request.POST.get("hrs_per_week") or 0
        course.credit = request.POST.get("credit") or 0
        course.marks_cia = request.POST.get("marks_cia") or 0
        course.marks_ese = request.POST.get("marks_ese") or 0
        course.total_marks = request.POST.get("total_marks") or 0
        course.is_saved = False
        course.is_finalized = False
        course.save()

        if old_course_code and old_course_code != new_course_code:
            CourseContent.objects.filter(course_code=old_course_code).update(course_code=new_course_code)

        messages.success(request, f"Course {course.course_code} updated successfully. Please save and confirm it again.")
        return redirect('upload_center:upload_center')

    return render(request, "add_course.html", {
        **_program_form_context(_course_form_initial(course)),
        "form_mode": "edit",
        "course": course,
    })


def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Course Template"

    headings = [
        'year', 'prog_type', 'prog_category', 'sem', 'course_code',
        'part', 'course_category', 'course_title', 'hrs_per_week',
        'credit', 'marks_cia', 'marks_ese', 'total_marks'
    ]

    for col_num, heading in enumerate(headings, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = heading
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    if not request.user.is_authenticated:
        return redirect('/login/')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="course_upload_template.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='/login/')
def get_filter_options(request):
    filters = {field: request.GET.get(field) for field in UPLOAD_FILTER_FIELDS}
    queryset = CourseStr.objects.all()
    return JsonResponse(_build_upload_filter_options(queryset, filters))
