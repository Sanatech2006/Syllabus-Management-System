from django.contrib.auth import get_user_model
from io import BytesIO

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse

from .models import Program


class ProgramManageTests(TestCase):
    def setUp(self):
        user = get_user_model().objects.create_user(
            username="tester",
            password="secret123",
            is_superuser=True,
            is_staff=True,
        )
        self.client.force_login(user)

    def test_program_management_waits_until_all_filters_are_selected(self):
        Program.objects.create(degree="B.A English", prog_type="UG", prog_category="Arts", prog_code="BAENG", branch="English")
        Program.objects.create(degree="M.Sc Chemistry", prog_type="PG", prog_category="Science", prog_code="MSCHEM", branch="Chemistry")

        response = self.client.get(reverse("program_manage:program_management"), {"degree": "B.A English"})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["has_applied_filters"])
        self.assertEqual(response.context["page_obj"].paginator.count, 0)
        self.assertEqual(len(response.context["all_programs_for_search"]), 0)

    def test_program_management_filters_after_all_filters_are_selected(self):
        Program.objects.create(degree="B.A English", prog_type="UG", prog_category="Arts", prog_code="BAENG", branch="English")
        Program.objects.create(degree="M.Sc Chemistry", prog_type="PG", prog_category="Science", prog_code="MSCHEM", branch="Chemistry")

        response = self.client.get(
            reverse("program_manage:program_management"),
            {
                "prog_type": "UG",
                "prog_category": "Arts",
                "degree": "B.A English",
                "branch": "English",
                "prog_code": "BAENG",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["has_applied_filters"])
        self.assertEqual(response.context["page_obj"].paginator.count, 1)
        self.assertEqual(response.context["page_obj"].object_list[0].prog_code, "BAENG")
        self.assertEqual(len(response.context["all_programs_for_search"]), 1)

    def test_get_filter_options_limits_related_program_filters_by_degree(self):
        Program.objects.create(degree="B.A English", prog_type="UG", prog_category="Arts", prog_code="BAENG", branch="English")
        Program.objects.create(degree="B.A English", prog_type="PG", prog_category="Science", prog_code="MSCHEM", branch="Chemistry")
        Program.objects.create(degree="B.Sc Physics", prog_type="UG", prog_category="Science", prog_code="BSCPHY", branch="Physics")

        response = self.client.get(reverse("program_manage:get_filter_options"), {"degree": "B.A English"})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["prog_types"], ["PG", "UG"])
        self.assertEqual(data["prog_codes"], ["BAENG", "MSCHEM"])
        self.assertEqual(data["branches"], ["Chemistry", "English"])
        self.assertEqual(data["degrees"], ["B.A English", "B.Sc Physics"])

    def test_download_template_contains_program_columns(self):
        response = self.client.get(reverse("program_manage:download_sample"))

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [sheet.cell(row=1, column=index).value for index in range(1, 6)]
        self.assertEqual(headers, ["prog_type", "prog_category", "degree", "branch", "prog_code"])

    def test_bulk_upload_creates_programs_from_template_columns(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["prog_type", "prog_category", "degree", "branch", "prog_code"])
        sheet.append(["UG", "Arts", "B.A English", "English", "BAENG"])
        sheet.append(["PG", "Science", "M.Sc Chemistry", "Chemistry", "MSCHEM"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        upload = SimpleUploadedFile(
            "programs.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("program_manage:upload_programs"),
            {"excel_file": upload},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertTrue(Program.objects.filter(prog_code="BAENG", degree="B.A English").exists())
        self.assertTrue(Program.objects.filter(prog_code="MSCHEM", degree="M.Sc Chemistry").exists())
